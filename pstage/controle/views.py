from django.shortcuts import render, redirect, HttpResponse
from openpyxl import load_workbook
from .models import *
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required

def admindjango(request):
    return redirect('/admin/')


def administrateur(request):
    num_agence = Agence.objects.all().count
    num_agent = Agent.objects.all().count
    num_direction = Direction.objects.all().count
    num_fonction = Fonction.objects.all().count
    num_caisse = Caisse.objects.all().count
    num_service = Service.objects.all().count
    
    return render(request, 'controle/administrateur.html', {"num_agence":num_agence, "num_agent":num_agent, "num_direction":num_direction, "num_caisse":num_caisse, "num_fonction":num_fonction, "num_service":num_service})


def ccco(request):
    return render(request, 'controle/ccco.html')


def search_consolidations(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        if start_date and end_date:
            consolidations = Consolidation.objects.filter(date__range=[start_date, end_date])
        else:
            consolidations = Consolidation.objects.all()

    return render(request, 'controle/consolidation.html', {'consolidations': consolidations})


def consolidation(request):
    nature_correcte = Nature.objects.get(libelle='Correcte')
    nature_different = Nature.objects.get(libelle='Difference')
    # nature_inexistant = Nature.objects.get(libelle='Inexistant')
    # nature_montant = Nature.objects.get(libelle='Montant Différent')
    # nature_telephone = Nature.objects.get(libelle='Numero de télephone différent')
    consolidations = []

    operations = Operation.objects.all()
    bordereauops = BordereauOp.objects.all()
    agences = Agence.objects.all()
    operateurs = Operateur.objects.all()

    # Initialisez la consolidation à None
    
    
    for operation in operations:
        consolidation_created = False  # Pour suivre si une consolidation a été créée pour cette opération

        for bordereauop in bordereauops:
            if (
                operation.date == bordereauop.date and
                operation.montant == bordereauop.amount and
                operation.refOp == bordereauop.idtransfer[-6:] and
                operation.client.tel == bordereauop.statement[-8:]
                ##ajouter comparaison des numero les 8 derniers chiffre de statement 
                # T25088 56046089
            ):
                consolidation = Consolidation(
                    operation=operation,
                    bordereauOp=bordereauop,
                    nature=nature_correcte,
                    reglement=True,
                    date=operation.date,
                    datereglement=operation.date,
                    operateur=bordereauop.operateur, 
                    agence=bordereauop.agence
                )
                consolidation.save()
                consolidations.append(consolidation)
                consolidation_created = True
                break 

        if not consolidation_created:
            consolidation = Consolidation(
                operation=operation,
                bordereauOp=bordereauop,
                nature=nature_different, ### difference peut etre inexistant dans bordereau et presente dans operation vice versa
                reglement=False,
                date=operation.date,
                datereglement=operation.date,
                operateur=bordereauop.operateur, 
                agence=bordereauop.agence
            )
            consolidation.save()
            consolidations.append(consolidation)
       
    return render(request, "controle/consolidation.html", {"consolidations": consolidations})


def importer(request):
    agences = Agence.objects.all()
    operateurs = Operateur.objects.all()
    if request.method == 'POST':
        agence_id  = request.POST['agence']
        agence = Agence.objects.get(pk=agence_id)
        operateur_id  = request.POST['operateur']
        operateur = Operateur.objects.get(pk=operateur_id)
        f_operateur = request.FILES['f_operateur']
        if f_operateur.name.endswith('.xlsx'):
            wb = load_workbook(f_operateur, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=10, values_only=True):
                objet = BordereauOp()
                objet.agence = agence
                objet.operateur = operateur
                objet.date = row[0]
                objet.idtransfer = row[1]
                objet.msisdn = row[2]
                objet.number = row[3]
                objet.statement = row[4]
                objet.amount = row[5]
                objet.credit = row[6]
                objet.debit = row[7]
                objet.fees = row[8]
                objet.previous = row[9]
                objet.post = row[10]
                objet.save()
            messages.success(request, 'Importation réussie', extra_tags='success-message')  # Réponse en cas de réussite
        else:
            messages.error(request, 'le fichier doit etre sous format .xlsx', extra_tags='error-message')
    return render(request, 'controle/importer.html', {"agences":agences, "operateurs":operateurs})


def agent(request):
    num_client = Client.objects.all().count
    num_oper = Operation.objects.all().count
    return render(request, 'controle/agent.html', {"num_client":num_client, "num_oper":num_oper,})


def dpf(request):
    num_appro = Approvisionnement.objects.all().count
    return render(request, 'controle/dpf.html', {"num_appro":num_appro})


def ajouterAgence(request):
    if request.method == 'POST':
        NoAgence = request.POST['NoAgence']
        nom = request.POST['nom']
        tel = request.POST['tel']
        mail = request.POST['mail']
        localisation = request.POST['localisation']
        agence= Agence(NoAgence=NoAgence, nom=nom, tel=tel, mail=mail, localisation=localisation)
        agence.save()
        return redirect('listeAgence')
    else:
        return render(request, 'controle/ajouterAgence.html')


def listeAgence(request):
    agences = Agence.objects.all()
    return render(request, 'controle/listeAgence.html', {"agences":agences})


def EditAgence(request):
    return render(request, 'controle/editAgence.html')


def ajouterAgent(request):
    fonctions = Fonction.objects.all()
    if request.method == 'POST':
        matricule = request.POST['matricule']
        nom = request.POST['nom']
        prenom = request.POST['prenom']
        mail = request.POST['mail']
        tel = request.POST['tel']
        boitePostale = request.POST['boitePostale']
        fonction_id  = request.POST['fonction']
        fonction = Fonction.objects.get(pk=fonction_id)
        agent = Agent(matricule=matricule, nom=nom, prenom=prenom, mail=mail,
                      tel=tel, boitePostale=boitePostale, fonction=fonction)
        agent.save()
        return redirect('listeAgent')
    return render(request, 'controle/ajouterAgent.html', {"fonctions":fonctions})


def listeAgent(request):
    agents = Agent.objects.all()
    return render(request, 'controle/listeAgent.html', {"agents":agents})


def ajouterAppro(request):
    agences = Agence.objects.all()
    directions = Direction.objects.all()
    compteops = CompteOp.objects.all()
    if request.method == 'POST':
        montant = request.POST['montant']
        date = request.POST['date']
        agence_id  = request.POST['agence']
        agence = Agence.objects.get(pk=agence_id)
        direction_id  = request.POST['direction']
        direction = Direction.objects.get(pk=direction_id)
        compteOp_id  = request.POST['compteOp'] 
        compteOp  = CompteOp.objects.get(pk=compteOp_id) 
        appro = Approvisionnement(montant=montant, date=date, agence=agence, direction=direction, compteOp=compteOp)
        appro.save()
        return redirect('listeAppro')
    return render(request, 'controle/ajouterAppro.html', {"agences":agences, "compteops":compteops, "directions":directions})


def listeAppro(request):
    appros = Approvisionnement.objects.all()
    return render(request, 'controle/listeAppro.html', {"appros": appros})


def ajouterCaisse(request):
    agents = Agent.objects.all()
    agences = Agence.objects.all()
    if request.method == 'POST':
        noCaisse = request.POST['noCaisse']
        compte = request.POST['compte']
        agence_id  = request.POST['agence']
        agence = Agence.objects.get(pk=agence_id)
        agent_id  = request.POST['agent']
        agent = Agent.objects.get(pk=agent_id)
        caisse = Caisse(noCaisse=noCaisse, compte=compte, agence=agence, agent=agent)
        caisse.save()
        return redirect('listeCaisse')
    return render(request, 'controle/ajouterCaisse.html', {"agences":agences, "agents":agents})


def listeCaisse(request):
    caisses= Caisse.objects.all()
    return render(request, 'controle/listeCaisse.html', {"caisses": caisses})


def ajouterClient(request):
    if request.method == 'POST':
        nom = request.POST['nom']
        prenom = request.POST['prenom']
        refCNIB = request.POST['refCNIB']
        tel = request.POST['tel']
        client= Client(nom=nom, prenom=prenom, refCNIB=refCNIB, tel=tel)
        client.save()
        return redirect('listeClient')

    return render(request, 'controle/ajouterClient.html' )


def listeClient(request):
    clients = Client.objects.all()
    return render(request, 'controle/listeClient.html' , {'clients':clients})


def ajouterCompteOp(request):
    agences = Agence.objects.all()
    operateurs = Operateur.objects.all()
    if request.method == 'POST':
        compteop= request.POST['compteop']
        agence_id = request.POST['agence']
        agence = Agence.objects.get(pk=agence_id)
        operateur_id = request.POST['operateur']
        operateur = Operateur.objects.get(pk=operateur_id)
        compteOp= CompteOp(compteop=compteop , agence=agence, operateur=operateur)
        compteOp.save()
        return redirect('listeCompteOp')
    return render(request, 'controle/ajouterCompteOp.html', {"operateurs":operateurs, "agences":agences})


def listeCompteOp(request):
    compteops = CompteOp.objects.all()
    return render(request, 'controle/listeCompteOp.html', {"compteops":compteops})


def ajouterDirection(request):
    agents = Agent.objects.all()
    agences = Agence.objects.all()
    if request.method == 'POST':
        nom= request.POST['nom']
        tel= request.POST['tel']
        mail= request.POST['mail']
        agence_id = request.POST['agence']
        agence = Agence.objects.get(pk=agence_id)
        agent_id  = request.POST['agent']
        agent = Agent.objects.get(pk=agent_id)
        direction = Direction(nom=nom, tel=tel, mail=mail, agence=agence, agent=agent)
        direction.save()
        return redirect('listeDirection')
    return render(request, 'controle/ajouterDirection.html', {"agences":agences, "agents":agents})


def listeDirection(request):
    directions = Direction.objects.all()
    return render(request, 'controle/listeDirection.html', {"directions":directions})


def ajouterFonction(request):
    if request.method == 'POST':
        code = request.POST['code']
        libelle = request.POST['libelle']
        fonction = Fonction(code=code, libelle=libelle)
        fonction.save()
        return redirect('listeFonction')
    return render(request, 'controle/ajouterFonction.html')    


def listeFonction(request):
    fonctions = Fonction.objects.all()
    return render(request, 'controle/listeFonction.html', {"fonctions":fonctions})


def ajouterNature(request):
    if request.method == 'POST':
        libelle= request.POST['libelle']
        nature = Nature(libelle=libelle)
        nature.save()
        return redirect('listeNature')
    return render(request, 'controle/ajouterNature.html')


def listeNature(request):
    natures = Nature.objects.all()
    return render(request, 'controle/listeNature.html', {"natures":natures})


def ajouterOperateur(request):
    if request.method == 'POST':
        nom= request.POST['nom']
        tel= request.POST['tel']
        mail= request.POST['mail']
        operateur = Operateur(nom=nom, tel=tel, mail=mail)
        operateur.save()
        return redirect('listeOperateur')
    return render(request, 'controle/ajouterOperateur.html')


def listeOperateur(request):
    operateurs = Operateur.objects.all()
    return render(request, 'controle/listeOperateur.html', {"operateurs":operateurs})


def ajouterOperation(request):
    services = Service.objects.all()
    caisses = Caisse.objects.all()
    typeOps = TypeOperation.objects.all()
    clients = Client.objects.all()
    agences = Agence.objects.all()
    operateurs = Operateur.objects.all()
    agents = Agent.objects.all()
    if request.method == 'POST':
        date = request.POST['date']
        montant = request.POST['montant']
        refOp = request.POST['refOp']
        service_id  = request.POST['service']
        service = Service.objects.get(pk=service_id)
        caisse_id = request.POST['caisse']
        caisse = Caisse.objects.get(pk=caisse_id)
        agence_id  = request.POST['agence']
        agence = Agence.objects.get(pk=agence_id)
        typeOp_id  = request.POST['typeOp']
        typeOp = TypeOperation.objects.get(pk=typeOp_id)
        client_id  = request.POST['client']
        client = Client.objects.get(pk=client_id)
        operateur_id  = request.POST['operateur']
        operateur = Operateur.objects.get(pk=operateur_id)
        agent_id  = request.POST['agent']
        agent = Agent.objects.get(pk=agent_id)
        operation = Operation(date=date, montant=montant, refOp=refOp, service =service , agence =agence , typeOp =typeOp , client =client , caisse=caisse, operateur=operateur , agent=agent )
        operation.save()
        return redirect('listeOperation')
    return render(request, 'controle/ajouterOperation.html', {"agents":agents,"services":services, "agences":agences, "caisses":caisses,
                                                              "clients":clients, "typeOps":typeOps, "operateurs":operateurs})


def listeOperation(request):
    operations = Operation.objects.all()
    return render(request, 'controle/listeOperation.html', {"operations":operations})


def ajouterService(request):
    agences = Agence.objects.all()
    if request.method == 'POST':
        noService = request.POST['noService']
        nom = request.POST['nom']
        agence_id  = request.POST['agence']
        agence = Agence.objects.get(pk=agence_id)
        service = Service(noService=noService, nom=nom, agence=agence)
        service.save()
        return redirect('listeService')
    return render(request, 'controle/ajouterService.html', {"agences": agences})


def listeService(request):
    services = Service.objects.all()
    return render(request, 'controle/listeService.html', {"services":services})


def ajouterTypeOp(request):
    if request.method == 'POST':
        nom= request.POST['nom']
        typeOp = TypeOperation(nom=nom)
        typeOp.save()
        return redirect('listeTypeOp')
    return render(request, 'controle/ajouterTypeOp.html')


def listeTypeOp(request):
    typeOps = TypeOperation.objects.all()
    return render(request, 'controle/listeTypeOp.html', {"typeOps":typeOps})

   

